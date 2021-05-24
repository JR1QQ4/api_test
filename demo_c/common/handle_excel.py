#!/usr/bin/python
# -*- coding:utf-8 -*-
import os
from pprint import pprint

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

# excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "excel_test.xlsx")

# 操作 Excel 的 3 个对象：
#   工作簿（Workbook）
#   表单（Sheet）
#   单元格（Cell）

# 1. 获取 Workbook
# wb: Workbook = load_workbook(excel_path)
# print(wb)  # <openpyxl.workbook.workbook.Workbook object at 0x000001FE39B46348>
# 2. 获取 Sheet
# sh: Worksheet = wb['Cases']  # <Worksheet "Cases">
# print(sh)
# 2.1 获取最大行数和最大列数
# row_num = sh.max_row
# column_num = sh.max_column
# print(row_num, column_num)  # 7 4
# 2.2 遍历 Sheet
# for i in range(1, row_num + 1):
#     for j in range(1, column_num + 1):
#         print("row_i={}, column_j={}, cell_value={}".format(i, j, sh.cell(i, j).value))
#     print('*' * 60)
# 2.3 获取行
# rows = list(sh.rows)
# print(rows)
# print(rows[0][0])  # <Cell 'Cases'.A1>
# print(rows[0][0].value)  # 获取 A1 单元格的值
# 2.4 遍历的另一种写法
# for item in list(sh.rows)[1:]:  # 去除表头
#     for cell in item:
#         print(cell.value)
#     print("*" * 60)
# 2.5 获取表头
# titles = [cell.value for cell in list(sh.rows)[0]]  # 获取表头
# print(titles)
# 2.6 返回字典
# titles = [cell.value for cell in list(sh.rows)[0]]
# data_lists = []
# for item in list(sh.rows)[1:]:
#     temp_dict = {}
#     for i in range(len(item)):
#         temp_dict[titles[i]] = item[i].value
#     print(temp_dict)
#     data_lists.append(temp_dict)
# print(data_lists)
# 2.7 返回字典的另一种方式
# titles = [cell.value for cell in list(sh.rows)[0]]
# data_lists = []
# for item in list(sh.rows)[1:]:
#     temp = []
#     for cell in item:
#         temp.append(cell.value)
#     res = dict(zip(titles, temp))
#     res['Params'] = eval(res['Params'])
#     data_lists.append(res)
# print(data_lists)
# 3. 获取 Cell
# cell: Cell = sh.cell(row=1, column=1)
# print(cell.value)  # <Cell 'Cases'.A1>
# 3.1 设置 Cell
# sh.cell(8, 1, 'cell')  # 或者使用 sh.cell(8, 1).value = 'cell'
# 4. 保存修改
# wb.save(excel_path)
# 5. 关闭 workbook
# wb.close()
from demo_c.common.handle_path import sources_dir


class HandleExcel:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.wb: Workbook = load_workbook(file_path)
        self.sh: Worksheet = self.wb[sheet_name]

    def get_titles(self):
        titles = []
        for cell in list(self.sh.rows)[0]:
            titles.append(cell.value)
        return titles

    def get_all_cases(self):
        all_cases = []
        titles = self.get_titles()
        for row in list(self.sh.rows)[1:]:
            values = []
            for cell in row:
                values.append(cell.value)
            res = dict(zip(titles, values))
            res['request_data'] = eval(str(res['request_data']))
            res['expected'] = eval(str(res['expected']))
            all_cases.append(res)
        return all_cases

    def write_in_excel(self, row, column, value=None):
        self.sh.cell(row, column, value)
        self.wb.save(self.file_path)
        self.wb.close()
        self.wb = load_workbook(self.file_path)

    def close(self):
        self.wb.close()


if __name__ == '__main__':
    # excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "excel_test.xlsx")
    # exl = HandleExcel(excel_path, "Cases")
    # print(exl.get_titles())
    # print(exl.get_all_cases())

    excel_path2 = os.path.join(sources_dir, "weixin_api.xlsx")
    exl = HandleExcel(excel_path2, "创建成员")
    expected_index = exl.get_titles().index('expected')
    pprint(exl.get_all_cases(), indent=2)
    exl.write_in_excel(2, expected_index + 1, "abc")
    # excel_path2 = os.path.join(sources_dir, "weixin_api.xlsx")
    pprint(exl.get_all_cases(), indent=2)
