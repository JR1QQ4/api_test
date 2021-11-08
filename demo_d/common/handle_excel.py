#!/usr/bin/python
# -*- coding:utf-8 -*-
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook


class HandleExcel:
    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        wb: Workbook = openpyxl.load_workbook(self.filename)
        sh: Worksheet = wb[self.sheetname]
        res = list(sh.rows)
        title = [i.value for i in res[0]]
        cases = []
        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        return cases

    def write_data(self, row, column, value):
        wb: Workbook = openpyxl.load_workbook(self.filename)
        sh: Worksheet = wb[self.sheetname]
        sh.cell(row=row, column=column, value=value)
        wb.save(self.filename)
