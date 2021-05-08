#!/usr/bin/python
# -*- coding:utf-8 -*-
from openpyxl import load_workbook
from openpyxl.workbook import workbook
from openpyxl.worksheet import worksheet


class HandleExcel:
    _wb: workbook
    _sh: worksheet
    titles = []
    all_data = []

    def __init__(self, file_path, sheet_name=None):
        self._file_path = file_path
        self._wb = load_workbook(file_path)
        if sheet_name is None:
            self._sh = self._wb.worksheets[0]
        else:
            self._sh = self._wb[sheet_name]
        self._max_row = self._sh.max_row
        self._max_column = self._sh.max_column
        self.get_title()
        # self._sh.cell(row=1, column=1)

    def get_title(self):
        for column in list(self._sh.rows)[0]:  # 遍历第 1 行中每一列
            val: str = column.value
            if val.find('(') != -1:
                val = val[0:val.find('(')]
            self.titles.append(val)
        return self.titles

    def convert(self, value: str):
        try:
            value = eval(value)
            return value
        except NameError:
            return value
        except Exception:
            self.convert(str(value))

    def get_all_data(self):
        titles = self.get_title()
        for row in list(self._sh.rows)[1:]:  # 遍历数据行
            values = []
            for cell in row:  # 获取每一行的值
                values.append(cell.value)
            res = dict(zip(titles, values))
            if 'Params' in res:  # 将 JSON 字符串转换为字典对象
                res['Params'] = self.convert(res['Params'])
            if 'Expectation' in res:
                res['Expectation'] = self.convert(res['Expectation'])
            if 'Actuality' in res:
                res['Actuality'] = self.convert(res['Actuality'])
            self.all_data.append(res)
        return self.all_data

    def update_all_data(self, data: list):
        for row_index, row_value in enumerate(data):
            for k, v in row_value.items():
                column = self.titles.index(k)
                self._sh.cell(row_index + 2, column + 1, str(v))
        self._wb.save(self._file_path)

    def update(self, data, value):
        self._sh.cell(int(data['Id']) + 1, self.titles.index('Actuality') + 1, value=str(value))
        self._wb.save(self._file_path)

    def __del__(self):
        self._wb.close()


if __name__ == '__main__':
    from demo_b.util.handle_path import data_dir
    import os

    # exl = os.path.join(data_dir, 'weixin.xlsx')
    # he = HandleExcel(exl, '读取成员')
    # print(he.get_title())
    # print(he.get_all_data())

    # a = [{'Id': 2, 'Desc': '用例2', 'Method': None, 'Url': 'https://qyapi.weixin.qq.com/cgi-bin/user/get',
    #       'Params': {'access_token': '#ACCESS_TOKEN#', 'userid': '#USERID#'},
    #       'Expectation': {'errcode': 0, 'errmsg': 'ok', 'userid': '#USERID#'}, 'Actuality': None}]
    # he.update_all_data(a)

    # b = {'access_token': '#ACCESS_TOKEN#', 'userid': '#USERID#'}
    # print(type(str(b)))

    # c = ['Id', 'Desc', 'Method', 'Url', 'Params', 'Expectation', 'Actuality']
    # print(c.index('Url'))

    # d = {'Id': 2, 'Desc': '用例2', 'Method': None, 'Url': 'https://qyapi.weixin.qq.com/cgi-bin/user/get',
    #      'Params': {'access_token': '#ACCESS_TOKEN#', 'userid': '#USERID#'},
    #      'Expectation': {'errcode': 0, 'errmsg': 'ok', 'userid': '#USERID#'}, 'Actuality': None}
    # he.update(d, "adc")
    # print(he.get_all_data())

    # a = "asd"
    # b = "{'a':123, 'b':213}"
    # try:
    #     print(eval(a))
    # except NameError:
    #     print(type(a))

    exl2 = os.path.join(data_dir, 'user.xlsx')
    he2 = HandleExcel(exl2)
    all_data = he2.get_all_data()
    insert_data = []
    def convert(value):
        if value is None:
            return ""
        return value
    for item in all_data:
        item['params'] = str({
            'access_token': convert(item['access_token']),
            'userid': convert(item['userid']),
            'name': convert(item['name']),
            'department': convert(item['department']),
            'mobile': convert(item['mobile'])
        })
        insert_data.append(item)
    print(all_data)
    he2.update_all_data(insert_data)
